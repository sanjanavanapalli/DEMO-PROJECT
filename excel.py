# storing email and password information
import openpyxl

path = 'C:/SELENIUM FINAL ASSESMENT/DATA.xlsx'
workbook = openpyxl.load_workbook(path)


# def create_sheet():
#     sheet = workbook.create_sheet('Login')
#     print(sheet.title)
#
#     workbook.save(path)


# create_sheet()

def set_data(email):
    data = [('Email', 'Password'), (email, 'S@njana07')]
    sheet = workbook['Login']
    for item in data:
        sheet.append(item)
    workbook.save(path)


def set_emailid(email):
    sheet = workbook['Login']
    email_id = sheet.cell(row=2, column=1)
    email_id.value = email
    workbook.save(path)

def total_price():
    sheet = workbook['Login']


# set_data(email)
# set_emailid(email_id)
